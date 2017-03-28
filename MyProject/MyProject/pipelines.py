# -*- coding: utf-8 -*-
import scrapy
from scrapy.exceptions import DropItem
from openpyxl import load_workbook,Workbook
import time
import os

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html

class MyprojectPipeline(object):
    def __init__(self):
        self.items = set()                #定义一个集合，去除重复记录
        self.columns = [
            "AnnouncementTitle",
            "Source",
            "ReportingDate",
            "AnnouncementSort",
            "FilePath"
        ]


    def process_item(self, item, spider):
        basepath = spider.settings['BASE_PATH']
        today = time.strftime("%Y%m%d", time.localtime(int(time.time())))
        today1 = time.strftime("%Y-%m-%d", time.localtime(int(time.time())))
        savePath = basepath+"ExchangeSpider\\Data\\"+today1+"\\"+"exchangeSpider_" + today + ".xlsx"
        if os.path.exists(savePath):
            wb = load_workbook(savePath)
            ws = wb['Sheet1']
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(self.columns)  # 设置表头
        try:
            if item['AnnouncementPath'] in self.items:
                raise DropItem(item)
            else:
                ws.append([item['AnnouncementTitle'].decode('utf-8'),
                            item['AnnouncementFrom'].decode('utf-8'),
                            item['ReportingDate'].decode('utf-8'),
                            item['AnnouncementSort'].decode('utf-8'),
                            item['AnnouncementPath'].decode('utf-8')
                            ])
                self.items.add(item['AnnouncementPath'])
            wb.save(savePath)
        except Exception,e:
            print e.message
        finally:
            return item





